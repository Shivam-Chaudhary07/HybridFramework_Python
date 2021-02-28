import openpyxl

def getRowCount(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    return sheet.max_row

def getColCount(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    return sheet.max_column

def readData(file, sheetname, rownum, colonum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    return(sheet.cell(row=rownum, column=colonum)).value

def writeData(file, sheetname, rownum, colonum, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    sheet.cell(row=rownum, column=colonum).value = data
    workbook.save(file)

#Create a TC sheet and import this sheet in it